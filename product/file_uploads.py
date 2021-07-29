from django.core.files.storage import default_storage, FileSystemStorage
from openpyxl import load_workbook
from product.models import Products, ProductType
import os, math, csv



def checkIfExist(path):
    return os.path.isdir(path)



def createDirectory(path):
    os.mkdir(path)



def saveFileData(path=None, contents=None):
    fs = FileSystemStorage()
    # this must indicate the filename and extension to make the file type same as original
    default_storage.save(path, contents)



def createChunkXlsx(file, path, filename):
    wb = load_workbook(file)
    sheetname = "Sheet1"
    ws = wb[sheetname]

    total_contents = getNumberOfRows(file, "xlsx")
    if not total_contents:
        return returnResponse(False, "File doesn't have a content")

    max_content_of_csv = 10
    number_of_chunks = 1 if total_contents <= max_content_of_csv else math.ceil(total_contents / max_content_of_csv)

    headers=[]
    for cell in ws[1]:
        headers.append(cell.value)

    all_rows = []
    for row in ws.iter_rows(min_row=2):
        all_rows.append(','.join([str(cell.value) for cell in row]))

    # create chunk of csv
    for x in range(number_of_chunks):
        chunk_filename = "{filename}_chunk_{chunk_number}.csv".format(filename=filename, chunk_number=x)
        full_directory = os.path.join(path, chunk_filename)
        with open(full_directory, mode='w', newline='') as csv_writer:
            csv_writer = csv.writer(csv_writer, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            start = x*max_content_of_csv
            end = (x*max_content_of_csv) + max_content_of_csv
            rows_to_insert = all_rows[start:end]
            for i in rows_to_insert:
                line = i.split(',')
                row_to_test = [data != None and data != 'None' for data in line]
                if not all(row_to_test):
                    continue
                csv_writer.writerow(line)
    # end of create chumk

    return returnResponse(True, "File chunk created")
        


def createChunkCsv(file, path, filename, temp, temp_filename):
    file_to_temp = temp +"\\"+ temp_filename + ".csv"
    saveFileData(file_to_temp, file)

    total_contents = getNumberOfRows(file_to_temp, "csv")
    if not total_contents:
        os.remove(file_to_temp)
        return returnResponse(False, "File doesn't have a content")

    all_rows = list()

    with open(file_to_temp, 'r') as f:
        for row in f:
            all_rows.append(row)

    all_rows = all_rows[1:]
    total_rows = len(all_rows) - 1 # headers not included

    max_content_of_csv = 10
    number_of_chunks = 1 if total_rows <= max_content_of_csv else math.ceil(total_rows / max_content_of_csv)
    
    # create chunk of csv
    for x in range(number_of_chunks):
        chunk_filename = "{filename}_chunk_{chunk_number}.csv".format(filename=filename, chunk_number=x)
        full_directory = os.path.join(path, chunk_filename)
        with open(full_directory, mode='w', newline='') as csv_writer:
            csv_writer = csv.writer(csv_writer, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            start = x*max_content_of_csv
            end = (x*max_content_of_csv) + max_content_of_csv
            rows_to_insert = all_rows[start:end]
            for i in rows_to_insert:
                csv_writer.writerow([item.rstrip() for item in i.split(',')])
    # end create chunk
    os.remove(file_to_temp)

    return returnResponse(True, "File chunk created")



def readContents(path, list_of_chunks):
    for file in list_of_chunks:
        actual_path = path + "\\" + file
        list_of_products = []
        with open(actual_path, newline='') as csvfile:
            rows = csv.reader(csvfile, delimiter=',', quotechar='"')
            for row in rows:
                product_type = ProductType.objects.get(pk=int(row[2]))
                product = Products(product_name=row[0], product_price="{:.2f}".format(float(row[1])), product_type=product_type, quantity=row[3], product_photo=row[4])
                list_of_products.append(product)
        importBulkOfProduct(list_of_products)



def importBulkOfProduct(list):
    bulks = Products.objects.bulk_create(list)



def returnResponse(success = None, message = None):
    return {
        "success": success,
        "message": message,
    }



def getNumberOfRows(file, file_type):
    number_of_content = 0

    if file_type == 'xlsx':
        wb = load_workbook(file)
        sheetname = "Sheet1"
        ws = wb[sheetname]
        for row in ws.iter_rows(min_row=2):
            if not all([cell.value == None or cell.value == 'None' for cell in row]):
                number_of_content += 1
    else:
        number_of_content = -1 # this is to make sure that header is not included for content validation
        with open(file, 'r') as f:
            for row in f:
                if not all([cell == None for cell in row]):
                    number_of_content += 1

    return number_of_content